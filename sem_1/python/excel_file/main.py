import openpyxl
wb=openpyxl.load_workbook("C:\\Users\\piyush kumar\\Downloads\\global_superstore.xlsx")
# print(type(wb))
ws=wb.active
# print(ws)
# print(f"maximum row in worksheet {ws.max_row} and maximum column in worksheet {ws.max_column}")
# working with cell
# print(ws['A1'].value)
# the value of cell 1 is order_id
# getting all the column 
# values=[ws.cell(row=i,column=6).value for i in range(2, ws.max_row+1)]
# print(values)

'''
# vv
my_list = []
# ---------- Step 3: Read rows from Excel ----------
for value in ws.iter_rows(
        min_row=1,
        max_row=11,
        min_col=1,
        max_col=4,
        values_only=True):
    my_list.append(value)
for ele1, ele2, ele3, ele4 in my_list:
    # print(ele1, ele2, ele3, ele4)
    print("{:<20}{:<50}{:<50}{:<20}".format(ele1, ele2, ele3, ele4))
'''

x=10
def x1():
    print(x)
    x=4
x1()